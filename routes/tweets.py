from flask import Blueprint, jsonify, request, send_file
from services.db_service import run_query
from services.post_tweets import post_tweet 
import logging
import random
from fpdf import FPDF
import os

logging.basicConfig(level=logging.INFO)

tweets_bp = Blueprint("tweets", __name__)

@tweets_bp.route("/tweets", methods=["GET"])
def get_collected_tweets():
    query = "SELECT source_username, tweet_id, tweet_text, created_at FROM collected_tweets ORDER BY created_at DESC LIMIT 50"
    tweets = run_query(query, fetchall=True)
    if not tweets:
        return jsonify({"message": "No hay tweets recolectados"}), 404
    return jsonify([{"source_username": t[0], "tweet_id": t[1], "tweet_text": t[2], "created_at": t[3]} for t in tweets]), 200


@tweets_bp.route("/post_tweet", methods=["POST"])
def post_tweet_route():
    data = request.json
    user_id = data.get("user_id")
    tweet_text = data.get("tweet_text")

    if not user_id or not tweet_text:
        return jsonify({"error": "Faltan parámetros (user_id o tweet_text)"}), 400

    if len(tweet_text) > 280:
        return jsonify({"error": "El texto del tweet excede el límite de 280 caracteres"}), 400

    response, status_code = post_tweet(user_id, tweet_text)

    return jsonify(response), status_code


@tweets_bp.route("/get-all-tweets/<twitter_id>", methods=["GET"])
def get_all_tweets(twitter_id):
    query = f"""
        SELECT user_id, source_value, tweet_id, tweet_text, created_at, priority
        FROM collected_tweets 
        WHERE user_id = '{twitter_id}' 
        ORDER BY created_at DESC
    """
    tweets = run_query(query, fetchall=True)
    
    if not tweets:
        return jsonify({"message": "No hay tweets recolectados para este usuario"}), 404
    
    return jsonify([
        {"user_id": t[0], "source_value": [1], "tweet_id": t[2], "tweet_text": t[3], "created_at": t[4], "priority": t[5]} 
        for t in tweets
    ]), 200


@tweets_bp.route("/get-posted-tweets/<twitter_id>", methods=["GET"])
def get_posted_tweets(twitter_id):
    query = f"""
        SELECT id, user_id, tweet_text, created_at
        FROM posted_tweets 
        WHERE user_id = '{twitter_id}' 
        ORDER BY created_at DESC
    """
    tweets = run_query(query, fetchall=True)
    
    if not tweets:
        return jsonify({"message": "No available tweets."}), 404
    
    return jsonify([
        {"id": t[0], "user_id": t[1], "tweet_text": t[2], "created_at": t[3]} 
        for t in tweets
    ]), 200


@tweets_bp.route("/delete-tweet/<tweet_id>", methods=["DELETE"])
def delete_tweet(tweet_id):
    query = f"DELETE FROM collected_tweets WHERE tweet_id = '{tweet_id}'"
    run_query(query)
    return jsonify({"message": "Tweet eliminado exitosamente"}), 200


@tweets_bp.route("/edit-tweet/<tweet_id>", methods=["PUT"])
def edit_tweet(tweet_id):
    data = request.json
    new_text = data.get("tweet_text")

    if not new_text:
        return jsonify({"error": "Faltan parámetros (tweet_text)"}), 400

    query = f"UPDATE collected_tweets SET tweet_text = '{new_text}' WHERE tweet_id = '{tweet_id}'"
    run_query(query)
    return jsonify({"message": "Tweet actualizado exitosamente"}), 200


@tweets_bp.route("/add-tweet", methods=["POST"])
def add_tweet():
    data = request.json
    user_id = data.get("user_id")
    tweet_text = data.get("tweet_text")

    if not user_id or not tweet_text:
        return jsonify({"error": "Faltan parámetros (user_id o tweet_text)"}), 400

    if len(tweet_text) > 280:
        return jsonify({"error": "El texto del tweet excede el límite de 280 caracteres"}), 400

    query = f"INSERT INTO collected_tweets (user_id, tweet_id, source_value, tweet_text, created_at) VALUES ('{user_id}', {random.randint(10**17, 10**18 -1)},'', '{tweet_text}', NOW())"
    run_query(query)
    return jsonify({"message": "Tweet agregado exitosamente"}), 201


@tweets_bp.route("/generate-pdf", methods=["GET"])
def generate_pdf():
    user_id = request.args.get("user_id")
    print(user_id)
    
    if not user_id:
        return jsonify({"error": "Se requiere el user_id"}), 400

    query = f"""
    SELECT u.username, ct.tweet_text, ct.created_at 
    FROM collected_tweets ct
    INNER JOIN users u ON ct.user_id = u.id
    WHERE ct.user_id = {user_id}
    ORDER BY ct.created_at DESC
    """
    tweets = run_query(query, fetchall=True)

    print(tweets)
    if not tweets:
        return jsonify({"error": "No hay tweets para este usuario"}), 404

    username = tweets[0][0] if tweets else "Unknown User"

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(200, 10, f"Collected tweets from: {username}", ln=True, align="C")
    
    pdf.set_font("Arial", size=12)
    for username, tweet_text, created_at in tweets:
        pdf.multi_cell(0, 10, f" - @{username}: {tweet_text}\n")
        pdf.ln(5)

    pdf_filename = f"tweets_backup_{username}.pdf"
    pdf_path = os.path.join("/tmp", pdf_filename)
    pdf.output(pdf_path)

    return send_file(pdf_path, as_attachment=True, download_name=pdf_filename)


@tweets_bp.route("/update-priority/<tweet_id>", methods=["PUT"])
def update_tweet_priority(tweet_id):
    data = request.json
    new_priority = data.get("priority")

    if new_priority is None:
        return jsonify({"error": "Falta el parámetro 'priority'"}), 400

    try:
        new_priority = int(new_priority)
        if new_priority not in [1, 2, 3]:
            return jsonify({"error": "La prioridad debe ser 1, 2 o 3"}), 400
    except ValueError:
        return jsonify({"error": "La prioridad debe ser un número entero"}), 400

    query = f"""
        UPDATE collected_tweets
        SET priority = {new_priority}
        WHERE tweet_id = '{tweet_id}'
    """
    run_query(query)
    return jsonify({"message": f"Prioridad actualizada a {new_priority}"}), 200


@tweets_bp.route("/provider-source", methods=["GET"])
def get_provider_source():
    row = run_query("SELECT value FROM global_config WHERE id = 1", fetchone=True)
    current = (row[0] if row else "").strip().upper()
    if current not in ("RAPIDAPI", "TWITTERAPI.IO"):
        current = "RAPIDAPI"
    return jsonify({"value": current}), 200


@tweets_bp.route("/provider-source", methods=["PUT"])
def set_provider_source():
    data = request.get_json(silent=True) or {}
    value = str(data.get("value", "")).strip().upper()

    allowed = {"RAPIDAPI", "TWITTERAPI.IO"}
    if value not in allowed:
        return jsonify({"error": "Valor inválido, use RAPIDAPI o TWITTERAPI.IO"}), 400

    existing = run_query("SELECT 1 FROM global_config WHERE id = 1", fetchone=True)
    if existing:
        run_query(f"UPDATE global_config SET value = '{value}' WHERE id = 1")
    else:
        run_query(f"INSERT INTO global_config (id, value) VALUES (1, '{value}')")

    return jsonify({"value": value}), 200


# ─── Extracted Tweets Management ────────────────────────────────────────────

@tweets_bp.route("/extracted/summary", methods=["GET"])
def get_extracted_summary():
    """Returns a per-account summary of extracted tweet counts."""
    rows = run_query(
        "SELECT u.id, u.twitter_id, u.username, u.profile_pic, u.account_status, COUNT(ct.id) as tweet_count "
        "FROM users u LEFT JOIN collected_tweets ct ON ct.user_id = u.id "
        "GROUP BY u.id, u.twitter_id, u.username, u.profile_pic, u.account_status ORDER BY u.username"
    )
    if not rows:
        return jsonify([]), 200
    return jsonify([{
        "user_id": r[0], "twitter_id": r[1], "username": r[2],
        "profile_pic": r[3], "account_status": r[4], "tweet_count": r[5]
    } for r in rows]), 200


@tweets_bp.route("/extracted/<int:user_id>", methods=["GET"])
def get_extracted_for_user(user_id):
    """Get all extracted tweets for a specific user_id with pagination."""
    limit = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0, type=int)
    rows = run_query(
        f"SELECT tweet_id, source_value, tweet_text, created_at, priority "
        f"FROM collected_tweets WHERE user_id = {user_id} "
        f"ORDER BY created_at DESC LIMIT {limit} OFFSET {offset}"
    )
    total_row = run_query(f"SELECT COUNT(*) FROM collected_tweets WHERE user_id = {user_id}", fetchone=True)
    total = total_row[0] if total_row else 0
    tweets = []
    if rows:
        for r in rows:
            tweets.append({
                "tweet_id": r[0], "source_value": r[1],
                "tweet_text": r[2],
                "created_at": r[3].isoformat() if r[3] else None,
                "priority": r[4]
            })
    return jsonify({"tweets": tweets, "total": total}), 200


@tweets_bp.route("/extracted/<int:user_id>/clear", methods=["DELETE"])
def clear_extracted_for_user(user_id):
    """Clear all extracted tweets for a specific account."""
    try:
        run_query(f"DELETE FROM collected_tweets WHERE user_id = {user_id}")
        return jsonify({"message": "All extracted tweets cleared for this account."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@tweets_bp.route("/extracted/clear-all", methods=["DELETE"])
def clear_all_extracted():
    """Clear ALL extracted tweets across all accounts."""
    try:
        run_query("DELETE FROM collected_tweets")
        return jsonify({"message": "All extracted tweets cleared across all accounts."}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@tweets_bp.route("/extracted/<int:user_id>/upload", methods=["POST"])
def upload_tweets_for_user(user_id):
    """
    Upload tweets from Excel/CSV for a specific account.
    Expected Excel columns: tweet_text (required), source_username (optional), tweet_id (optional)
    """
    try:
        import openpyxl, csv, io
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "No file uploaded"}), 400

        filename = file.filename.lower()
        tweets_to_insert = []

        if filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                tweet_text = str(row_data.get("tweet_text") or "").strip()
                if tweet_text:
                    tweets_to_insert.append({
                        "tweet_text": tweet_text,
                        "source_username": str(row_data.get("source_username") or "manual_upload"),
                        "tweet_id": str(row_data.get("tweet_id") or random.randint(10**17, 10**18 - 1))
                    })
        elif filename.endswith(".csv"):
            content = file.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                tweet_text = str(row.get("tweet_text") or "").strip()
                if tweet_text:
                    tweets_to_insert.append({
                        "tweet_text": tweet_text,
                        "source_username": row.get("source_username") or "manual_upload",
                        "tweet_id": row.get("tweet_id") or str(random.randint(10**17, 10**18 - 1))
                    })
        else:
            return jsonify({"error": "Only .xlsx and .csv files supported"}), 400

        inserted = 0
        for t in tweets_to_insert:
            txt = t["tweet_text"].replace("'", "''")
            src = str(t["source_username"]).replace("'", "")
            tid = str(t["tweet_id"]).replace("'", "")
            try:
                run_query(
                    f"INSERT INTO collected_tweets (user_id, tweet_id, source_value, tweet_text, created_at) "
                    f"VALUES ({user_id}, '{tid}', '{src}', '{txt}', NOW()) ON CONFLICT DO NOTHING"
                )
                inserted += 1
            except Exception:
                pass

        return jsonify({"message": f"Uploaded {inserted} tweets", "count": inserted}), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@tweets_bp.route("/extract/manual/<string:twitter_id>", methods=["POST"])
def manual_extract_for_account(twitter_id):
    """Trigger a manual tweet extraction for one specific account."""
    try:
        from services.fetch_tweets import fetch_tweets_for_single_user
        user_row = run_query(f"SELECT id FROM users WHERE twitter_id = '{twitter_id}'", fetchone=True)
        if not user_row:
            return jsonify({"error": "Account not found"}), 404
        user_id = user_row[0]
        import threading
        stop_event = threading.Event()
        t = threading.Thread(target=lambda: fetch_tweets_for_single_user(user_id, stop_event), daemon=True)
        t.start()
        return jsonify({"message": f"Manual extraction started for @{twitter_id}"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@tweets_bp.route("/extract/manual/all", methods=["POST"])
def manual_extract_all():
    """Trigger a manual tweet extraction for all active accounts."""
    try:
        from services.fetch_tweets import fetch_tweets_for_all_users
        import threading
        stop_event = threading.Event()
        t = threading.Thread(target=lambda: fetch_tweets_for_all_users(stop_event), daemon=True)
        t.start()
        return jsonify({"message": "Manual extraction started for all active accounts"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Posted Tweets Management ────────────────────────────────────────────────

@tweets_bp.route("/posted/summary", methods=["GET"])
def get_posted_summary():
    """Returns a per-account summary of posted tweet counts."""
    rows = run_query(
        "SELECT u.id, u.twitter_id, u.username, u.profile_pic, u.account_status, COUNT(pt.id) as posted_count "
        "FROM users u LEFT JOIN posted_tweets pt ON pt.user_id = u.id "
        "GROUP BY u.id, u.twitter_id, u.username, u.profile_pic, u.account_status ORDER BY u.username"
    )
    if not rows:
        return jsonify([]), 200
    return jsonify([{
        "user_id": r[0], "twitter_id": r[1], "username": r[2],
        "profile_pic": r[3], "account_status": r[4], "posted_count": r[5]
    } for r in rows]), 200


@tweets_bp.route("/posted/<int:user_id>/by-day", methods=["GET"])
def get_posted_by_day(user_id):
    """Returns posted tweets grouped by day for a specific user."""
    rows = run_query(
        f"SELECT DATE(created_at) as post_date, COUNT(*) as count "
        f"FROM posted_tweets WHERE user_id = {user_id} "
        f"GROUP BY DATE(created_at) ORDER BY post_date DESC LIMIT 30"
    )
    days = []
    if rows:
        for r in rows:
            days.append({"date": str(r[0]), "count": r[1]})
    return jsonify(days), 200


@tweets_bp.route("/posted/<int:user_id>/on-date", methods=["GET"])
def get_posted_on_date(user_id):
    """Returns all posts for a user on a specific date."""
    target_date = request.args.get("date")  # e.g. 2026-07-16
    if not target_date:
        return jsonify({"error": "date query param required (YYYY-MM-DD)"}), 400
    rows = run_query(
        f"SELECT id, tweet_text, created_at, post_status, failure_reason "
        f"FROM posted_tweets WHERE user_id = {user_id} AND DATE(created_at) = '{target_date}' "
        f"ORDER BY created_at ASC"
    )
    tweets = []
    if rows:
        for r in rows:
            tweets.append({
                "id": r[0], "tweet_text": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "status": r[3] or "posted",
                "failure_reason": r[4]
            })
    return jsonify(tweets), 200

